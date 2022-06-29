
from main_oop import RecordingTransactions
from openpyxl import load_workbook

wb_tl = load_workbook("UYUSMA ROBOTU/TL.xlsx")
ws_tl = wb_tl.active
recorder = RecordingTransactions()
max_row_tl = ws_tl.max_row

for row in ws_tl.iter_rows(min_row=17, max_row=max_row_tl - 7):
    if row[0].value is None:
        pass
    else:

        just_date = row[0].value.split("-")[0]
        just_time = row[0].value.split("-")[1]
        recorder.adding_transactions(just_date, row[3].value, row[0].coordinate, row[8].value)
print(recorder.print_out_days())

