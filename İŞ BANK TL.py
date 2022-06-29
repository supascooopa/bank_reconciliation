#BİSMİLLAHİRAHMANİRAHİM

import openpyxl
import datetime
from pprint import pprint
import dateutil.parser
from collections import defaultdict

wb = openpyxl.open("UYUSMA ROBOTU/680475124_31.08.2021_10 58 10_HesapOzeti.xlsx")
ws = wb.active
date_column = ws["A"] #Tarih sütunu
transaction_amount_column = ws["D"] #İşlem tutarı sütunu
transaction_type_column = ws["H"] #İşlem tipi sütunu

date_from = datetime.datetime.now()
# max_day = int(input("Please enter the number of days you would like to go back: "))

main_dict = defaultdict(list)

for row in ws.iter_rows(min_row=17):
    if row[0].value is None:
        pass
    else:

        just_date = row[0].value.split("-")[0]
        print(just_date)
        main_dict[just_date].extend([[row[0].row, row[3].value, row[7].value]])


# TODO 1.: FIND THE SAME TRANSACTIONS IN ONE DAY
# TODO 2.: SEPERATE TRANSACTIONS TYPES
# TODO 3.: WRITE THEM TO A NEW EXCEL SHEET
# TODO 4.: FILL OUT THE NEW EXCEL SHEET WITH THE CORRESPONDING ACCOUNTING SERIAL NUMBERS AND MARKINGS
pprint(main_dict)
print(main_dict['31/08/2021'][0][0])
