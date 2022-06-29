import openpyxl
from openpyxl import load_workbook, Workbook
import datetime
from pprint import pprint
from collections import defaultdict
from itertools import zip_longest


def writer(counter, amounts, accounting_code_b, accounting_code_a, transaction_desc, foreign_currency):
    counts = counter
    new_ws["H4"] = f"Tarih: {going_back_formatted}"
    new_ws[f"a{counts}"] = accounting_code_b
    new_ws[f"g{counts}"] = foreign_currency
    new_ws[f"h{counts}"] = amounts
    new_ws[f"d{counts}"] = transaction_desc
    new_ws[f"f{counts}"] = "B"
    counts += 1
    new_ws[f"a{counts}"] = accounting_code_a
    new_ws[f"h{counts}"] = amounts
    new_ws[f"d{counts}"] = transaction_desc
    new_ws[f"f{counts}"] = "A"
    counts += 1
    new_wb.save(f"{going_back_formatted_microsoft}.xlsx")
    return counts


wb_tl = load_workbook("UYUSMA ROBOTU/TL.xlsx")
ws_tl = wb_tl.active
wb_usd = load_workbook("UYUSMA ROBOTU/USD.xlsx")
ws_usd = wb_usd.active
wb_eur = load_workbook("UYUSMA ROBOTU/EUR.xlsx")
ws_eur = wb_eur.active
now = datetime.datetime.now()

transaction_dict_tl = defaultdict(list)
transaction_dict_usd = defaultdict(list)
transaction_dict_eur = defaultdict(list)
max_row_tl = ws_tl.max_row
max_row_usd = ws_usd.max_row
max_row_eur = ws_eur.max_row
for row in ws_tl.iter_rows(min_row=17, max_row=max_row_tl - 7):
    if row[0].value is None:
        pass
    else:

        just_date = row[0].value.split("-")[0]
        just_time = row[0].value.split("-")[1]
        transaction_dict_tl[just_date].extend([[row[3].value, row[7].value, row[8].value, just_time]])

for row in ws_usd.iter_rows(min_row=17, max_row=max_row_usd - 7):
    if row[0].value is None:
        pass
    else:
        just_date = row[0].value.split("-")[0]
        just_time = row[0].value.split("-")[1]
        transaction_dict_usd[just_date].extend([[row[3].value, row[6].value, row[7].value, just_time]])

for row in ws_eur.iter_rows(min_row=17, max_row=max_row_eur - 7):
    if row[0].value is None:
        pass
    else:

        just_date = row[0].value.split("-")[0]
        just_time = row[0].value.split("-")[1]
        transaction_dict_eur[just_date].extend([[row[3].value, row[6].value, row[7].value, just_time]])

now = datetime.datetime.now()
# choice = int(input("please input how many day you want to go back to: "))
choice = 10

for date in range(1, choice):
    going_back = now - datetime.timedelta(date)
    going_back_formatted = going_back.strftime("%d/%m/%Y")
    going_back_formatted_microsoft = going_back.strftime("%d.%m.%Y")
    if not transaction_dict_tl[going_back_formatted]:
        pass
    else:
        new_wb = openpyxl.load_workbook("UYUSMA ROBOTU/BLUECELL YEVMIYE ORNEGI.xlsx")
        new_ws = new_wb.active
        count = 8

        for tl in transaction_dict_tl[going_back_formatted]:
            amount_tl = tl[0]
            transaction_type_tl = tl[1]
            description_tl = tl[2]
            time_of_transaction_tl = tl[3]
            #POS AKTARIMLARI
            if transaction_type_tl == "POS" and amount_tl > 0:
                count = writer(counter=count, amounts=amount_tl, accounting_code_b="1-0-2-00-130",
                               accounting_code_a="1-0-8-30-400", transaction_desc="POS AKT", foreign_currency=0)
            #DOVIZ ALIMLARI
            #1.USD
            if "6804-0061802 İŞCEP DÖVİZ ALIŞ" in description_tl:
                for usd in transaction_dict_usd[going_back_formatted]:
                    amount_usd = usd[0]
                    transaction_type_usd = usd[1]
                    description_usd = usd[2]
                    time_of_transaction_usd = usd[3]
                    if "6804-0075124 İŞCEP DÖVİZ ALIŞ" in description_usd and\
                            time_of_transaction_usd == time_of_transaction_tl:

                        count = writer(counter=count, amounts=abs(amount_tl), accounting_code_b="1-0-2-05-130",
                                       accounting_code_a="1-0-2-00-130", transaction_desc="USD AKT",
                                       foreign_currency=amount_usd)
            #2.EUR
            if "6804-0061817 İNTERNETTEN DÖVİZ ALIŞ" in description_tl:
                for eur in transaction_dict_eur[going_back_formatted]:
                    amount_eur = eur[0]
                    transaction_type_eur = eur[1]
                    description_eur = eur[2]
                    time_of_transaction_eur = eur[3]
                    if "6804-0075124 İNTERNETTEN DÖVİZ ALIŞ" in description_eur and\
                            time_of_transaction_eur == time_of_transaction_tl:

                        count = writer(counter=count, amounts=abs(amount_tl), accounting_code_b="1-0-2-05-130",
                                       accounting_code_a="1-0-2-05-131", transaction_desc="EUR AKT",
                                       foreign_currency=amount_eur)
            #BANKA MASRAFLARI
            if "DVZ ALIŞ KMV TUT" in description_tl:
                count = writer(counter=count, amounts=abs(amount_tl), accounting_code_b="7-9-4-48-480",
                               accounting_code_a="1-0-2-00-130", transaction_desc="BANKA MASRAFI", foreign_currency=0)

            elif transaction_type_tl == "POS" and amount_tl < 0:
                count = writer(counter=count, amounts=abs(amount_tl), accounting_code_b="7-9-4-48-480",
                               accounting_code_a="1-0-2-00-130", transaction_desc="BANKA MASRAFI", foreign_currency=0)

            elif transaction_type_tl == "Havale" and "HAV. ÜZ." in description_tl:
                count = writer(counter=count, amounts=abs(amount_tl), accounting_code_b="7-9-4-48-480",
                               accounting_code_a="1-0-2-00-130", transaction_desc="BANKA MASRAFI", foreign_currency=0)
            elif "4508********0019 İŞCEP KRE.KART BORÇ ÖDEME" in description_tl:
                count = writer(counter=count, amounts=abs(amount_tl), accounting_code_b="3-0-0-00-130",
                               accounting_code_a="1-0-2-00-130", transaction_desc="0019 İLE BİTEN KK BORÇ ÖDEMESİ",
                               foreign_currency=0)
            elif "4508********2584 İŞCEP KRE.KART BORÇ ÖDEME" in description_tl:
                count = writer(counter=count, amounts=abs(amount_tl), accounting_code_b="3-0-0-00-131",
                               accounting_code_a="1-0-2-00-130", transaction_desc="2584 İLE BİTEN KK BORÇ ÖDEMESİ",
                               foreign_currency=0)



