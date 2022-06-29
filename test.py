
from openpyxl import load_workbook, Workbook
import datetime

from collections import defaultdict


def pos(counter, amounts, accounting_code_b, accounting_code_a, transaction_desc):
    counts = counter
    new_ws["J1"] = going_back_formatted
    new_ws[f"a{counts}"] = accounting_code_b
    new_ws[f"h{counts}"] = amounts
    new_ws[f"d{counts}"] = transaction_desc
    new_ws[f"f{counts}"] = "B"
    counts += 1
    new_ws[f"a{counts}"] = accounting_code_a
    new_ws[f"h{counts}"] = amounts
    new_ws[f"d{counts}"] = transaction_desc
    new_ws[f"f{counts}"] = "A"
    counts += 1
    return counts


wb_tl = load_workbook("UYUSMA ROBOTU/bu hafta Lira.xlsx")
ws_tl = wb_tl.active
wb_usd = load_workbook("UYUSMA ROBOTU/bu hafta dolar.xlsx")
ws_usd = wb_usd.active
now = datetime.datetime.now()

transaction_dict_tl = defaultdict(list)
transaction_dict_usd = defaultdict(list)
max_row = ws_tl.max_row
for row in ws_tl.iter_rows(min_row=17, max_row=max_row - 7):
    if row[0].value is None:
        pass
    else:
        just_date = row[0].value.split("-")[0]
        transaction_dict_tl[just_date].extend([[row[3].value, row[7].value, row[8].value]])
for row in ws_usd.iter_rows(min_row=17, max_row=max_row - 7):
    if row[0].value is None:
        pass
    else:
        just_date = row[0].value.split("-")[0]
        transaction_dict_usd[just_date].extend([[row[3].value, row[7].value, row[8].value]])
now = datetime.datetime.now()
# choice = int(input("please input how many day you want to go back to: "))
choice = 10

new_wb = Workbook()
new_ws = new_wb.active

for date in range(1, choice):
    going_back = now - datetime.timedelta(date)
    going_back_formatted = going_back.strftime("%d/%m/%Y")
    going_back_formatted_microsoft = going_back.strftime("%d.%m.%Y")
    #TODO 3.: Need to find the date in USD dict
    if not transaction_dict_tl[going_back_formatted]:
        pass
    else:
        new_wb = Workbook()
        new_ws = new_wb.active
        count = 1
        #TODO 4.: Need to find the correct list in USD dict
        for x in transaction_dict_tl[going_back_formatted]:
            amount = x[0]
            transaction_type = x[1]
            description = x[2]
            print(transaction_type)
            if x[1] == "POS" and x[0] > 0:
                count = pos(counter=count, amounts=amount, accounting_code_b="1-0-2-00-130",
                            accounting_code_a="1-0-8-30-400", transaction_desc="POS AKT")
            #     print(count)
            # print(f"total: {count}")
            # new_wb.save(f"{going_back_formatted_microsoft}.xlsx")
            if amount < 0 and transaction_type == "Döviz" and "İŞCEP DÖVİZ ALIŞ" in description:

                count = pos(counter=count, amounts=abs(amount), accounting_code_b="1-0-2-05-130",
                            accounting_code_a="1-0-2-00-130", transaction_desc="USD AKT")
            #     #TODO 2.: Need to do the USD side of this transaction, then we are golden!

            if "DVZ ALIŞ KMV TUT" in description:
                count = pos(counter=count, amounts=abs(amount), accounting_code_b="7-9-4-48-480",
                            accounting_code_a="1-0-2-00-130", transaction_desc="BANKA MASRAFI")
            #     #TODO: Do a function for the below sequence since it is being repeated many times!


